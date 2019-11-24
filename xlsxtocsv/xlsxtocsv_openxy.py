from openpyxl import load_workbook
import pandas as pd


wb = load_workbook(filename='D:\\gitfiles\\xlsxtocsv.xlsx')
sheetnames = wb.get_sheet_names()

for x in sheetnames:
  data_xls = pd.read_excel('D:\\gitfiles\\xlsxtocsv.xlsx', x, index_col=None)
  data_xls.to_csv( 'D:\\gitfiles\\'+ x +'.csv', encoding='utf-8', index=False)  