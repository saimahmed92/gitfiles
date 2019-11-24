from openpyxl import load_workbook
import openpyxl
import csv

wb = load_workbook(filename='D:\\gitfiles\\xlsxtocsv.xlsx')

sheetnames = wb.get_sheet_names()
sheet_value_arr = []


for a in sheetnames:
 sheet = wb[a]
 with open('D:\\example_'+ a +'.csv', 'w') as f:
    c = csv.writer(f)
    for row in sheet.rows:
        sheet_value_arr.append([cell.value for cell in row])
        for r in sheet.rows:
            c.writerow([cell.value for cell in row])

f.close()